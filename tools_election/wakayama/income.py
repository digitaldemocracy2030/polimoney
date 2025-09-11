import re

from openpyxl.worksheet.worksheet import Worksheet

from util import A_COL, B_COL, C_COL, E_COL, J_COL, extract_number


def get_individual_income(income: Worksheet):
    """収入の部の個別データを取得する

    A列 = 0, B列 = 1, ...
    結合されているセルは、左端のセルに情報が入っている

    Args:
        income (Worksheet): 収入の部のシート

    Returns:
        list: 収入の部のデータ
    """
    income_data = []

    # 4行目以降, AからJの列を取得
    min_row = 4

    for row in income.iter_rows(min_row=min_row, max_col=J_COL + 1):
        date_cell = row[A_COL]
        price_cell = row[C_COL]
        category_cell = row[E_COL]
        note_cell = row[J_COL]
        # Noneになったら終了
        if date_cell.value is None:
            break

        income_data.append(
            {
                "date": date_cell.value.strftime("%Y-%m-%d"),
                "price": extract_number(price_cell.value),
                "category": category_cell.value,
                "note": note_cell.value,
            }
        )

    return income_data


def get_total_income(income: Worksheet):
    """総収入を取得する
    合計に関する記述は9行あり、位置は個別データの数によって変わるので、動的に取得する

    Args:
        income (Worksheet): 収入の部のシート
    """

    total_income_data = []
    count = 0

    # 合計に関する記述は7行目より下にある
    min_row = 7

    for row in income.iter_rows(min_row=min_row, max_col=C_COL + 1):
        # B列が寄附, その他の収入, 計, 総計のいずれでもなければスキップ
        if row[B_COL].value not in ["寄附", "その他の収入", "計", "総計"]:
            continue
        # 9行取得したら終了
        if count == 9:
            break

        name_value = row[B_COL].value
        price_value = row[C_COL].value

        price_value = extract_number(price_value)
        total_income_data.append({"name": name_value, "price": price_value})
        count += 1

    return total_income_data


def get_public_expense_equivalent(income: Worksheet):
    """公費負担相当額を取得する
    位置は個別データの数によって変わるので、動的に取得する
    ここだけ文章で記述されているため、正規表現でパースして辞書形式で返す

    Args:
        income (Worksheet): _description_
    """

    # 公費負担相当額に関する記述は7 + 9 + 2 = 18行目より下にある
    min_row = 18
    public_expense_equivalent_str = ""

    # ここではBしか取得しないため、row[0]で取得できる
    for row in income.iter_rows(min_row=min_row, max_col=B_COL + 1):
        if row[A_COL].value == "参考":
            public_expense_equivalent_str = row[B_COL].value
            break
    else:
        return {}

    # 総額を取得する正規表現
    total_pattern = r"公費負担相当額\s+(\d+(?:,\d+)*)円"
    total_match = re.search(total_pattern, public_expense_equivalent_str)

    # 内訳を取得する正規表現
    breakdown_pattern = r"内訳\s+(.+)"
    breakdown_match = re.search(breakdown_pattern, public_expense_equivalent_str)

    public_expense_equivalent_data = {}

    # 総額を追加
    if total_match:
        total_amount = int(total_match.group(1).replace(",", ""))
        public_expense_equivalent_data["total"] = total_amount

    # 内訳をパース
    if breakdown_match:
        breakdown_str = breakdown_match.group(1)
        # カンマを事前に削除
        breakdown_str = breakdown_str.replace(",", "")

        # 内訳内の各項目をパースする正規表現（項目名 + 数字 + 円 の形式）
        # スペースが無くても対応するため、数字の直前までを項目名とする
        item_pattern = r"([^0-9]+)(\d+)円"
        items = re.findall(item_pattern, breakdown_str)

        breakdown_data = {}
        for item_name, amount_str in items:
            # 項目名の前後のスペース、カンマ、読点を除去
            item_name = item_name.strip().replace("、", "").replace(",", "")
            # 金額を整数に変換
            amount = int(amount_str)
            breakdown_data[item_name] = amount

        public_expense_equivalent_data["breakdown"] = breakdown_data

    return public_expense_equivalent_data


def get_income(income: Worksheet):
    individual_income = get_individual_income(income)

    total_income = get_total_income(income)

    public_expense_equivalent = get_public_expense_equivalent(income)

    return {
        "individual_income": individual_income,
        "total_income": total_income,
        "public_expense_equivalent": public_expense_equivalent,
    }
