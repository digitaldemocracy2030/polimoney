import datetime

from openpyxl.worksheet.worksheet import Worksheet

from util import A_COL, B_COL, C_COL, E_COL, F_COL, K_COL, extract_number


def get_individual_election_office(building: Worksheet):
    """家屋費(選挙事務所費)の個別データを取得する

    Args:
        building (Worksheet): 人件の部のシート
    """

    building_data = []

    min_row = 4
    for row in building.iter_rows(min_row=min_row, max_col=K_COL + 1):
        date_cell = row[A_COL]
        price_cell = row[C_COL]
        category_cell = row[E_COL]
        purpose_cell = row[F_COL]
        note_cell = row[K_COL]
        # Noneになったら終了
        if date_cell.value is None:
            break

        building_data.append(
            {
                "date": date_cell.value.strftime("%Y-%m-%d"),
                "price": extract_number(price_cell.value),
                "category": category_cell.value,
                "purpose": purpose_cell.value,
                "note": note_cell.value,
            }
        )

    return building_data


def get_total_election_office(building: Worksheet):
    """家屋費(選挙事務所費)の合計データを取得する
    合計に関する記述は3行あり、位置は個別データの数によって変わるので、動的に取得する

    Args:
        building (Worksheet): 家屋の部のシート
    """

    total_building_data = []
    count = 0
    json_checksum = 0  # jsonファイルの検証に使用

    # 合計に関する記述は16行目より下にある
    min_row = 16

    for row in building.iter_rows(min_row=min_row, max_col=C_COL + 1):
        # 型をチェック
        value = row[B_COL].value
        if not isinstance(value, str):
            continue

        # B列が立候補準備のための支出、選挙運動のための支出、計のいずれでもなければスキップ
        value_str = value.strip().replace("　", "")
        if value_str not in ["立候補準備のための支出", "選挙運動のための支出", "計"]:
            continue

        # 3行取得したら終了
        if count == 3:
            break

        price_value = row[C_COL].value
        price_value = extract_number(price_value)
        total_building_data.append({"name": value_str, "price": price_value})
        count += 1
        if value_str == "計":
            json_checksum = price_value

    return total_building_data, json_checksum


def get_individual_meeting_venue(building: Worksheet):
    """集会会場費の個別データを取得する

    Args:
        building (Worksheet): 集会会場費のシート

    Returns:
        list: 集会会場費の個別データのリスト
    """
    meeting_venue_data = []

    # 22行目以降からスタートする スタート位置を特定
    min_row = 20
    start_row = 0

    for i, row in enumerate(
        building.iter_rows(min_row=min_row, max_col=A_COL + 1), start=min_row
    ):
        date_cell = row[A_COL]
        if date_cell.value is not None and date_cell.value == "月　　日":
            start_row = i + 2
            break
    else:
        return []

    # 特定した行以降からスタートする
    for row in building.iter_rows(min_row=start_row, max_col=K_COL + 1):
        # 型をチェック
        date_cell = row[A_COL]
        if not isinstance(date_cell.value, datetime.datetime):
            continue

        date_cell = row[A_COL]
        price_cell = row[C_COL]
        category_cell = row[E_COL]
        purpose_cell = row[F_COL]
        note_cell = row[K_COL]
        # Noneになったら終了
        if date_cell.value is None:
            break

        meeting_venue_data.append(
            {
                "date": date_cell.value.strftime("%Y-%m-%d"),
                "price": extract_number(price_cell.value),
                "category": category_cell.value,
                "purpose": purpose_cell.value,
                "note": note_cell.value,
            }
        )

    return meeting_venue_data


def get_total_meeting_venue(building: Worksheet):
    """集会会場費の合計データを取得する
    合計に関する記述は3行あり、位置は個別データの数によって変わるので、動的に取得する

    Args:
        building (Worksheet): 集会会場費のシート
    """

    total_meeting_venue_data = []
    count = 0
    json_checksum = 0  # jsonファイルの検証に使用

    # 合計に関する記述は34行目より下にある
    min_row = 34

    for row in building.iter_rows(min_row=min_row, max_col=C_COL + 1):
        # 型をチェック
        value = row[B_COL].value
        if not isinstance(value, str):
            continue

        # B列が立候補準備のための支出、選挙運動のための支出、計のいずれでもなければスキップ
        value_str = value.strip().replace("　", "")
        if value_str not in ["立候補準備のための支出", "選挙運動のための支出", "計"]:
            continue

        # 3行取得したら終了
        if count == 3:
            break

        price_value = row[C_COL].value
        price_value = extract_number(price_value)
        total_meeting_venue_data.append({"name": value_str, "price": price_value})
        count += 1
        if value_str == "計":
            json_checksum = price_value

    return total_meeting_venue_data, json_checksum


def get_building(building: Worksheet):
    individual_election_office = get_individual_election_office(building)

    total_election_office, json_checksum1 = get_total_election_office(building)

    individual_meeting_venue = get_individual_meeting_venue(building)

    total_meeting_venue, json_checksum2 = get_total_meeting_venue(building)

    json_checksum = json_checksum1 + json_checksum2

    return {
        "individual_election_office": individual_election_office,
        "total_election_office": total_election_office,
        "individual_meeting_venue": individual_meeting_venue,
        "total_meeting_venue": total_meeting_venue,
        "json_checksum": json_checksum,
    }
