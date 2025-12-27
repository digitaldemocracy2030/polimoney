import logging
import sys

import openpyxl

import util
from wakayama.building import get_building
from wakayama.general import get_general
from wakayama.income import get_income
from wakayama.summary import get_summary

# ログ設定
logging.basicConfig(
    level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s"
)


"""

和歌山県は、入力ファイルが複数の場合、すべてのファイルが必要なため、
最初に定数としてファイル名を定義。

"""

INPUT_FILES = {
    1: "",  # 第1回
    2: "",  # 第2回
    3: "",  # 第3回
}


def analyze(file_path):
    """
    指定されたExcelファイルを解析し、各シートのデータをJSONファイルとして出力する。

    人件・通信・交通・広告・文具・食料・休泊・雑費は同じフォーマットで処理される。

    Args:
        file_path (str): 解析対象のExcelファイルのパス
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)

    # 各シートを取得
    income = wb["収入"]
    personnel = wb["人件"]
    building = wb["家屋"]
    communication = wb["通信"]
    transportation = wb["交通"]
    printing = wb["印刷"]
    advertising = wb["広告"]
    stationery = wb["文具"]
    food = wb["食料"]
    accommodation = wb["休泊"]
    miscellaneous = wb["雑費"]
    summary = wb["支出 (計)"]

    # 分析
    income_data = get_income(income)  # 収入
    personnel_data = get_general(personnel, "personnel")  # 人件
    building_data = get_building(building)  # 家屋
    communication_data = get_general(communication, "communication")  # 通信
    transportation_data = get_general(transportation, "transportation")  # 交通
    printing_data = get_general(printing, "printing")  # 印刷
    advertising_data = get_general(advertising, "advertising")  # 広告
    stationery_data = get_general(stationery, "stationery")  # 文具
    food_data = get_general(food, "food")  # 食料
    accommodation_data = get_general(accommodation, "accommodation")  # 休泊
    miscellaneous_data = get_general(miscellaneous, "miscellaneous")  # 雑費
    summary_data = get_summary(summary)  # 支出計

    # フォルダを作成
    safe_input_file = util.create_output_folder(file_path)

    # データとファイル名を定義 (utilで必要)
    data_list = [
        ("income_data.json", income_data),
        ("personnel_data.json", personnel_data),
        ("building_data.json", building_data),
        ("communication_data.json", communication_data),
        ("transportation_data.json", transportation_data),
        ("printing_data.json", printing_data),
        ("advertising_data.json", advertising_data),
        ("stationery_data.json", stationery_data),
        ("food_data.json", food_data),
        ("accommodation_data.json", accommodation_data),
        ("miscellaneous_data.json", miscellaneous_data),
        ("summary_data.json", summary_data),
    ]

    util.create_individual_json(data_list, safe_input_file)

    return data_list


def analyze_income(income_file_path: str, key: int):
    """
    収入データは、一番最初に提出されたファイルから解析を行う。
    収支報告書Excelファイルが複数あり、かつ収入と支出のデータが分かれている場合に使用する。

    Args:
        income_file_path (str): 収入データが含まれているExcelファイルのパス
        combined_file_path (str): 結合データのパス
    """
    wb = openpyxl.load_workbook(income_file_path, data_only=True)

    # 各シートを取得
    income = wb["収入"]
    income_data = get_income(income)

    # フォルダを作成
    safe_input_file = util.create_output_folder(income_file_path)

    # 収入データのみを個別データとして出力
    data_list = [
        (f"income_data_{key}.json", income_data),
    ]
    util.create_individual_json(data_list, safe_input_file)

    return data_list


def analyze_communication(file_path: str, key: int):
    """
    通信費データはすべてのファイルにそれぞれユニークなデータがあるため、
    個別に解析してJSONファイルとして出力する。

    Args:
        file_path (str): 解析対象のExcelファイルのパス
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)

    # 通信シートを取得
    communication = wb["通信"]
    communication_data = get_general(communication, "communication")

    # フォルダを作成
    safe_input_file = util.create_output_folder(file_path)

    # 通信データのみを個別データとして出力
    data_list = [
        (f"communication_data_{key}.json", communication_data),
    ]
    util.create_individual_json(data_list, safe_input_file)

    return data_list


def main():
    """メイン関数。コマンドライン引数から入力ファイルを取得し、解析処理を実行する。

    Returns:
        int: 正常終了時は0を返す。

    Raises:
        SystemExit: コマンドライン引数が不正な場合、エラーメッセージを表示して終了する。
    """
    logging.info(f"分析を開始します: {INPUT_FILES.values()}")

    data_list = []

    # 最初に提出されたファイルを解析
    first_file_path = INPUT_FILES[min(INPUT_FILES.keys())]
    data = analyze_income(first_file_path, 1)
    data_list.extend(data)

    # 最後以外のすべてのファイルを解析
    for key in sorted(INPUT_FILES.keys()):
        if key == max(INPUT_FILES.keys()):
            continue
        file_path = INPUT_FILES[key]
        data = analyze_communication(file_path, key)
        data_list.extend(data)

    # 最後に提出されたファイルを解析
    last_file_path = INPUT_FILES[max(INPUT_FILES.keys())]
    data = analyze(last_file_path)
    data_list.extend(data)

    output_folder = util.create_output_folder(last_file_path)

    util.create_combined_json(data_list, output_folder)

    logging.info(f"分析を完了しました: {INPUT_FILES.values()}")

    return 0


if __name__ == "__main__":
    sys.exit(main())
